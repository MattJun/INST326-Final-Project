
import openpyxl
from openpyxl import Workbook

class ScheduleMaker:
    def __init__(self, input_filename = "Schedule.xlsx"):
        self.dataframe = openpyxl.load_workbook("Schedule.xlsx")
        self.dataframe1 = self.dataframe.active
        self.wb = Workbook()
        print("creating schedule object")
    

    def create_output_sheet(self):
        self.sheet1 = self.wb.create_sheet(index=0, title="schedule")
        self.wb.save("May_8.xlsx")
        print("creating output")

    def write_header(self):
        self.sheet1 = self.wb.create_sheet(index = 0, title = "Week Schedule")
        self.sheet1.cell(row = 1, column=2).value = "Monday"
        self.sheet1.cell(row = 1, column =3).value = "Tuesday"
        self.sheet1.cell(row =1, column =4).value = "Wednesday"
        self.sheet1.cell(row=1, column=5).value = "Thursday"
        self.sheet1.cell(row =1, column =6).value = "Friday"
        print("wrote headers")

    def write_time_slots(self):
        self.sheet1.cell(row = 2, column = 1).value = "Times"
        self.sheet1.cell(row =3, column =1).value = "9:00"
        self.sheet1.cell(row =4, column =1).value = "10:00"
        self.sheet1.cell(row =5, column =1).value = "11:00"
        self.sheet1.cell(row =6, column =1).value = "12:00"
        self.sheet1.cell(row =7, column =1).value = "1:00"
        self.sheet1.cell(row =8, column =1).value = "2:00"
        self.sheet1.cell(row =9, column =1).value = "3:00"
        self.sheet1.cell(row =10, column =1).value = "4:00"
        self.sheet1.cell(row =11, column =1).value = "5:00"
        print("wrote time slots")

    


    def write_schedule(self):
        for i in range(2, self.dataframe1.max_row + 1):
            for j in range(3, self.dataframe1.max_column + 1):
                stg = str(self.dataframe1.cell(i, j).value)
                list = stg.split("-")
                row_num = 3
                name = self.dataframe1.cell(i, 1).value
                for x in list:
                  row_num = 3
                  #checking 9:00am    
                  if x == "9":
                    if self.dataframe1.cell(i, 4).value == "Y":
                        self.sheet1.cell(row_num, 2).value = name
                    if self.dataframe1.cell(i, 5).value == "Y":
                        self.sheet1.cell(row_num, 3).value = name
                    if self.dataframe1.cell(i, 6).value == "Y":
                        self.sheet1.cell(row_num, 4).value = name
                    if self.dataframe1.cell(i, 7).value == "Y":
                        self.sheet1.cell(row_num, 5).value = name
                    if self.dataframe1.cell(i, 8).value == "Y":
                        self.sheet1.cell(row_num, 6).value = name
                  row_num += 1
                  #checking 10:00am
                  if x == "10":
                    if self.dataframe1.cell(i, 4).value == "Y":
                        self.sheet1.cell(row_num, 2).value = name
                    if self.dataframe1.cell(i, 5).value == "Y":
                        self.sheet1.cell(row_num, 3).value = name
                    if self.dataframe1.cell(i, 6).value == "Y":
                        self.sheet1.cell(row_num, 4).value = name
                    if self.dataframe1.cell(i, 7).value == "Y":
                        self.sheet1.cell(row_num, 5).value = name
                    if self.dataframe1.cell(i, 8).value == "Y":
                        self.sheet1.cell(row_num, 6).value = name      
                  row_num += 1
                  #checking 11:00am
                  if x == "11":
                    if self.dataframe1.cell(i, 4).value == "Y":
                        self.sheet1.cell(row_num, 2).value = name
                    if self.dataframe1.cell(i, 5).value == "Y":
                        self.sheet1.cell(row_num, 3).value = name
                    if self.dataframe1.cell(i, 6).value == "Y":
                        self.sheet1.cell(row_num, 4).value = name
                    if self.dataframe1.cell(i, 7).value == "Y":
                        self.sheet1.cell(row_num, 5).value = name
                    if self.dataframe1.cell(i, 8).value == "Y":
                        self.sheet1.cell(row_num, 6).value = name
                  row_num += 1
                  #Checking 12:00am      
                  if x == "12":
                    if self.dataframe1.cell(i, 4).value == "Y":
                        self.sheet1.cell(row_num, 2).value = name
                    if self.dataframe1.cell(i, 5).value == "Y":
                        self.sheet1.cell(row_num, 3).value = name
                    if self.dataframe1.cell(i, 6).value == "Y":
                        self.sheet1.cell(row_num, 4).value = name
                    if self.dataframe1.cell(i, 7).value == "Y":
                        self.sheet1.cell(row_num, 5).value = name
                    if self.dataframe1.cell(i, 8).value == "Y":
                        self.sheet1.cell(row_num, 6).value = name
                  row_num += 1
                  #checking 1:00pm
                  if x == "1":
                    if self.dataframe1.cell(i, 4).value == "Y":
                        self.sheet1.cell(row_num, 2).value = name
                    if self.dataframe1.cell(i, 5).value == "Y":
                        self.sheet1.cell(row_num, 3).value = name
                    if self.dataframe1.cell(i, 6).value == "Y":
                        self.sheet1.cell(row_num, 4).value = name
                    if self.dataframe1.cell(i, 7).value == "Y":
                        self.sheet1.cell(row_num, 5).value = name
                    if self.dataframe1.cell(i, 8).value == "Y":
                        self.sheet1.cell(row_num, 6).value = name
                  row_num += 1
                  #checking 2:00pm
                  if x == "2":
                    if self.dataframe1.cell(i, 4).value == "Y":
                        self.sheet1.cell(row_num, 2).value = name
                    if self.dataframe1.cell(i, 5).value == "Y":
                        self.sheet1.cell(row_num, 3).value = name
                    if self.dataframe1.cell(i, 6).value == "Y":
                        self.sheet1.cell(row_num, 4).value = name
                    if self.dataframe1.cell(i, 7).value == "Y":
                        self.sheet1.cell(row_num, 5).value = name
                    if self.dataframe1.cell(i, 8).value == "Y":
                        self.sheet1.cell(row_num, 6).value = name
                  row_num += 1
                  #checking 3:00pm
                  if x == "3":
                    if self.dataframe1.cell(i, 4).value == "Y":
                        self.sheet1.cell(row_num, 2).value = name
                    if self.dataframe1.cell(i, 5).value == "Y":
                        self.sheet1.cell(row_num, 3).value = name
                    if self.dataframe1.cell(i, 6).value == "Y":
                        self.sheet1.cell(row_num, 4).value = name
                    if self.dataframe1.cell(i, 7).value == "Y":
                        self.sheet1.cell(row_num, 5).value = name
                    if self.dataframe1.cell(i, 8).value == "Y":
                        self.sheet1.cell(row_num, 6).value = name        
                  row_num += 1
                  #checking 4:00pm
                  if x == "4":
                    if self.dataframe1.cell(i, 4).value == "Y":
                        self.sheet1.cell(row_num, 2).value = name
                    if self.dataframe1.cell(i, 5).value == "Y":
                        self.sheet1.cell(row_num, 3).value = name
                    if self.dataframe1.cell(i, 6).value == "Y":
                        self.sheet1.cell(row_num, 4).value = name
                    if self.dataframe1.cell(i, 7).value == "Y":
                        self.sheet1.cell(row_num, 5).value = name
                    if self.dataframe1.cell(i, 8).value == "Y":
                        self.sheet1.cell(row_num, 6).value = name
                  row_num += 1
                  #checking 5:00pm
                  if x == "5":
                    if self.dataframe1.cell(i, 4).value == "Y":
                        self.sheet1.cell(row_num, 2).value = name
                    if self.dataframe1.cell(i, 5).value == "Y":
                        self.sheet1.cell(row_num, 3).value = name
                    if self.dataframe1.cell(i, 6).value == "Y":
                        self.sheet1.cell(row_num, 4).value = name
                    if self.dataframe1.cell(i, 7).value == "Y":
                        self.sheet1.cell(row_num, 5).value = name
                    if self.dataframe1.cell(i, 8).value == "Y":
                        self.sheet1.cell(row_num, 6).value = name
    print("wrote schedule")
# def getworkbook(self):                       
#     filename = "may_8.xlsx"
#     self.wb.save(filename)
#     print(filename)   
                                         
       


def main():
    currSchedule = ScheduleMaker("Schedule.xlsx")
    currSchedule.write_header()
    currSchedule.create_output_sheet()
    currSchedule.write_schedule() 
    currSchedule.wb.save("May_8.xlsx")
    

if __name__ == "__main__":
    print(main())
    
    