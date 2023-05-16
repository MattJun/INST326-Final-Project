import openpyxl
from openpyxl import Workbook


"""This class is used to create people to add to the schedule if 
the person creating the schedule wants to add someone """

class People:
    
    """__init__ initializes filename"""
    
    def __init__(self, filename = "Schedule.xlsx", userAnswer = ""): #Driver:Dan  Navigator: Arianna
        self.filename = filename
        self.userAnswer = userAnswer

    """add_person function checks if the workbook exists, 
    then writes in the workbook. This function then prompts the user in the console
    if they would like to add someone to the schedule, asking what time they can work
    along with asking which day they are available, marking a Y or N depending if that person can work or not.
    We also added an edge case if someone accidently types something wrong in the Console, 
    to make sure the prompt is responded to correctly
    
    -Purpose is for business owner or schedule maker to update the schedule without the need
    to go to the excelsheet to do it manually"""

    def add_person(self, userAnswer):
        try:
            workbook = openpyxl.load_workbook(self.filename)
        except FileNotFoundError:
            workbook = openpyxl.Workbook("Schedule.xlsx")

        sheet = workbook.active
        availability = []
        days_of_week = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']  
        
        if userAnswer != "yes" and userAnswer != "no":
            print("would you like to add someone to the schedule?")
            userAnswer = input("yes or no: ")
        
        while userAnswer != "no" and userAnswer != "yes":
            userAnswer = input("""Please answer with "yes" or "no": """)
        if userAnswer == "yes":    
            name = input("Enter Name: ")
            number = input("Time he/she can work:")
        
            for day in days_of_week:
                answer = input(f"Is {name} available on {day}? (Y/N): ")
                while answer != "N" and answer != "Y":
                    answer = input(f"Please answer with(Y/N). Is {name} available on {day}?: ")
                availability.append("Y" if answer.upper() == 'Y' else '')

            sheet.append([name,"", number] + availability)

            workbook.save(self.filename)
            workbook.close()
            print(f"Person '{name}' added to the Excel sheet '{self.filename}'.")
            
            userAnswer = input("Would you like to add again?: ")
            while userAnswer != "yes" and userAnswer != "no":
                userAnswer = input("""Please respond with "yes" or "no". Would you like to add again?: """)
            if userAnswer == "yes":
                people.add_person("yes")
            elif userAnswer =="no":
                print("Thank you for using!")
        elif userAnswer == "no":
            print("Thank you for using!")

people = People("Schedule.xlsx")
people.add_person("")
            



class ScheduleMaker:
    
    """This __init__ function initalizes dataframe, dataframe1, wb, and sheet1"""
    
    def __init__(self, input_filename="Schedule.xlsx"): #Driver: Arianna  Navigator: Jennifer
        self.dataframe = openpyxl.load_workbook("Schedule.xlsx")
        self.dataframe1 = self.dataframe.active
        self.wb = Workbook()
        self.sheet1 = self.wb.active
        print("Creating ScheduleMaker object")

    """The function create_output_sheet creates a sheet within the workbook for the schedule to be written"""

    def create_output_sheet(self): #Driver: Dan  Navigator: Matt
        self.sheet1 = self.wb.create_sheet(index=0, title="schedule")
        self.wb.save("May_11.xlsx")
        print("Creating output sheet")

    """The write_header function is used to write down the header of the schedule, 
    having the days of the week layed out on the top row."""

    def write_header(self): #Driver: Arianna  Navigator: Dan
        self.sheet1.cell(row=1, column=2).value = "Monday"
        self.sheet1.cell(row=1, column=3).value = "Tuesday"
        self.sheet1.cell(row=1, column=4).value = "Wednesday"
        self.sheet1.cell(row=1, column=5).value = "Thursday"
        self.sheet1.cell(row=1, column=6).value = "Friday"
        print("Wrote headers")

    """The write_time_slots function is used to go down the row and write the different times
    an employee can work for the schedule."""

    def write_time_slots(self): #Driver: Matt  Navigator: Dan
        self.sheet1.cell(row=2, column=1).value = "Times"
        self.sheet1.cell(row=3, column=1).value = "9:00"
        self.sheet1.cell(row=4, column=1).value = "10:00"
        self.sheet1.cell(row=5, column=1).value = "11:00"
        self.sheet1.cell(row=6, column=1).value = "12:00"
        self.sheet1.cell(row=7, column=1).value = "1:00"
        self.sheet1.cell(row=8, column=1).value = "2:00"
        self.sheet1.cell(row=9, column=1).value = "3:00"
        self.sheet1.cell(row=10, column=1).value = "4:00"
        self.sheet1.cell(row=11, column=1).value = "5:00"
        print("Wrote time slots")
    
    """This function is used to read the availability sheet, and then determine if the employee put a certain time, and a "Y" on
    the availability sheet, then with that information write it on the the new schedule, determining when that employee can work """

    def write_schedule(self): #Driver:Matt  Navigator: Jennifer
        for i in range(2, self.dataframe1.max_row + 1):
            for j in range(3, self.dataframe1.max_column + 1):
                stg = str(self.dataframe1.cell(i, j).value) # The i, j is the location in the excel sheet. i is the row, and j is the column.
                list = stg.split(",")
                row_num = 3
                name = self.dataframe1.cell(i, 1).value
                for x in list: #the x is the times for each person. 
                  row_num = 3
                  #checking 9:00am    
                  if x == "9":
                    if self.dataframe1.cell(i, 4).value == "Y":
                        self.sheet1.cell(row_num, 2).value = name
                    if self.dataframe1.cell(i, 5).value == "Y":
                        self.sheet1.cell(row_num, 3).value = name
                    if self.dataframe1.cell(i, 6).value == "Y":
                        self.sheet1.cell(row_num, 4).value = name
                    if self.dataframe1.cell(i, 7).value == "Y":
                        self.sheet1.cell(row_num, 5).value = name
                    if self.dataframe1.cell(i, 8).value == "Y":
                        self.sheet1.cell(row_num, 6).value = name
                  row_num += 1
                  #checking 10:00am
                  if x == "10":
                    if self.dataframe1.cell(i, 4).value == "Y":
                        self.sheet1.cell(row_num, 2).value = name
                    if self.dataframe1.cell(i, 5).value == "Y":
                        self.sheet1.cell(row_num, 3).value = name
                    if self.dataframe1.cell(i, 6).value == "Y":
                        self.sheet1.cell(row_num, 4).value = name
                    if self.dataframe1.cell(i, 7).value == "Y":
                        self.sheet1.cell(row_num, 5).value = name
                    if self.dataframe1.cell(i, 8).value == "Y":
                        self.sheet1.cell(row_num, 6).value = name      
                  row_num += 1
                  #checking 11:00am
                  if x == "11":
                    if self.dataframe1.cell(i, 4).value == "Y":
                        self.sheet1.cell(row_num, 2).value = name
                    if self.dataframe1.cell(i, 5).value == "Y":
                        self.sheet1.cell(row_num, 3).value = name
                    if self.dataframe1.cell(i, 6).value == "Y":
                        self.sheet1.cell(row_num, 4).value = name
                    if self.dataframe1.cell(i, 7).value == "Y":
                        self.sheet1.cell(row_num, 5).value = name
                    if self.dataframe1.cell(i, 8).value == "Y":
                        self.sheet1.cell(row_num, 6).value = name
                  row_num += 1
                  #Checking 12:00am      
                  if x == "12":
                    if self.dataframe1.cell(i, 4).value == "Y":
                        self.sheet1.cell(row_num, 2).value = name
                    if self.dataframe1.cell(i, 5).value == "Y":
                        self.sheet1.cell(row_num, 3).value = name
                    if self.dataframe1.cell(i, 6).value == "Y":
                        self.sheet1.cell(row_num, 4).value = name
                    if self.dataframe1.cell(i, 7).value == "Y":
                        self.sheet1.cell(row_num, 5).value = name
                    if self.dataframe1.cell(i, 8).value == "Y":
                        self.sheet1.cell(row_num, 6).value = name
                  row_num += 1
                  #checking 1:00pm
                  if x == "1":
                    if self.dataframe1.cell(i, 4).value == "Y":
                        self.sheet1.cell(row_num, 2).value = name
                    if self.dataframe1.cell(i, 5).value == "Y":
                        self.sheet1.cell(row_num, 3).value = name
                    if self.dataframe1.cell(i, 6).value == "Y":
                        self.sheet1.cell(row_num, 4).value = name
                    if self.dataframe1.cell(i, 7).value == "Y":
                        self.sheet1.cell(row_num, 5).value = name
                    if self.dataframe1.cell(i, 8).value == "Y":
                        self.sheet1.cell(row_num, 6).value = name
                  row_num += 1
                  #checking 2:00pm
                  if x == "2":
                    if self.dataframe1.cell(i, 4).value == "Y":
                        self.sheet1.cell(row_num, 2).value = name
                    if self.dataframe1.cell(i, 5).value == "Y":
                        self.sheet1.cell(row_num, 3).value = name
                    if self.dataframe1.cell(i, 6).value == "Y":
                        self.sheet1.cell(row_num, 4).value = name
                    if self.dataframe1.cell(i, 7).value == "Y":
                        self.sheet1.cell(row_num, 5).value = name
                    if self.dataframe1.cell(i, 8).value == "Y":
                        self.sheet1.cell(row_num, 6).value = name
                  row_num += 1
                  #checking 3:00pm
                  if x == "3":
                    if self.dataframe1.cell(i, 4).value == "Y":
                        self.sheet1.cell(row_num, 2).value = name
                    if self.dataframe1.cell(i, 5).value == "Y":
                        self.sheet1.cell(row_num, 3).value = name
                    if self.dataframe1.cell(i, 6).value == "Y":
                        self.sheet1.cell(row_num, 4).value = name
                    if self.dataframe1.cell(i, 7).value == "Y":
                        self.sheet1.cell(row_num, 5).value = name
                    if self.dataframe1.cell(i, 8).value == "Y":
                        self.sheet1.cell(row_num, 6).value = name        
                  row_num += 1
                  #checking 4:00pm
                  if x == "4":
                    if self.dataframe1.cell(i, 4).value == "Y":
                        self.sheet1.cell(row_num, 2).value = name
                    if self.dataframe1.cell(i, 5).value == "Y":
                        self.sheet1.cell(row_num, 3).value = name
                    if self.dataframe1.cell(i, 6).value == "Y":
                        self.sheet1.cell(row_num, 4).value = name
                    if self.dataframe1.cell(i, 7).value == "Y":
                        self.sheet1.cell(row_num, 5).value = name
                    if self.dataframe1.cell(i, 8).value == "Y":
                        self.sheet1.cell(row_num, 6).value = name
                  row_num += 1
                  #checking 5:00pm
                  if x == "5":
                    if self.dataframe1.cell(i, 4).value == "Y":
                        self.sheet1.cell(row_num, 2).value = name
                    if self.dataframe1.cell(i, 5).value == "Y":
                        self.sheet1.cell(row_num, 3).value = name
                    if self.dataframe1.cell(i, 6).value == "Y":
                        self.sheet1.cell(row_num, 4).value = name
                    if self.dataframe1.cell(i, 7).value == "Y":
                        self.sheet1.cell(row_num, 5).value = name
                    if self.dataframe1.cell(i, 8).value == "Y":
                        self.sheet1.cell(row_num, 6).value = name
    print("wrote schedule")
                   
       
"""This main function is used to individually run each function within each of the classes. 
Then save the schedule on the user's device"""

def main(): #Driver: Jennifer  Navigator:Arianna
    currSchedule = ScheduleMaker("Schedule.xlsx")
    currSchedule.create_output_sheet()
    currSchedule.write_header()
    currSchedule.write_time_slots()
    currSchedule.write_schedule() 
    currSchedule.wb.save("May_11.xlsx")
 
"""The __name == "__main__" is used to run the main funtion that runs the entire code. """

if __name__ == "__main__":
    main()
